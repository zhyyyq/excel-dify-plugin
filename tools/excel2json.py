import json
from collections.abc import Generator
from typing import Any
import urllib.request
import tempfile
import os

from dify_plugin import Tool
from dify_plugin.entities.tool import ToolInvokeMessage

from openpyxl import load_workbook


class Excel2JsonTool(Tool):
    def _invoke(self, tool_parameters: dict[str, Any]) -> Generator[ToolInvokeMessage]:
        file_meta = tool_parameters['file']
        
        # Get file path - could be local or remote URL
        # file_meta is a File object with url attribute
        file_url = file_meta.url
        if not file_url:
            raise Exception("No file URL provided")

        try:
            # Create temp file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                tmp_path = tmp.name
            
            # Download file if it's a remote URL
            if file_url.startswith('http://') or file_url.startswith('https://'):
                urllib.request.urlretrieve(file_url, tmp_path)
            else:
                # It's already a local path
                tmp_path = file_url

            # Load workbook to get data
            wb = load_workbook(tmp_path, data_only=True)
            sheet_names = wb.sheetnames

            # Load workbook to get styles
            wb_styles = load_workbook(tmp_path)

            if len(sheet_names) > 1:
                # Multiple sheets
                json_output = {}

                for sheet_name in sheet_names:
                    ws = wb[sheet_name]
                    data = self._read_sheet_data(ws)

                    # Extract styles
                    styles = {}
                    if sheet_name in wb_styles.sheetnames:
                        ws_styles = wb_styles[sheet_name]
                        styles = self._extract_sheet_styles(ws_styles)

                    # Extract dimensions (row heights and column widths)
                    dimensions = self._extract_sheet_dimensions(ws)

                    if styles:
                        json_output["[styles]"] = styles

                    if dimensions:
                        json_output["[format]"] = {"sheets": {sheet_name: dimensions}}

                    json_output[sheet_name] = data

                wb.close()
                wb_styles.close()
                
                # Clean up temp file if we created one
                if file_url.startswith('http://') or file_url.startswith('https://'):
                    os.unlink(tmp_path)
                
                yield self.create_text_message(json.dumps(json_output, ensure_ascii=False, indent=2))
            else:
                # Single sheet
                ws = wb[sheet_names[0]]
                data = self._read_sheet_data(ws)

                # Extract styles
                styles = {}
                if sheet_names[0] in wb_styles.sheetnames:
                    ws_styles = wb_styles[sheet_names[0]]
                    styles = self._extract_sheet_styles(ws_styles)

                # Extract dimensions (row heights and column widths)
                dimensions = self._extract_sheet_dimensions(ws)

                # Build output
                output = {sheet_names[0]: data}
                if dimensions:
                    output["[format]"] = {"sheets": {sheet_names[0]: dimensions}}
                if styles:
                    output["[styles]"] = styles

                wb.close()
                wb_styles.close()
                
                # Clean up temp file if we created one
                if file_url.startswith('http://') or file_url.startswith('https://'):
                    os.unlink(tmp_path)
                
                yield self.create_text_message(json.dumps(output, ensure_ascii=False, indent=2))

        except Exception as e:
            # Clean up temp file on error
            try:
                if 'tmp_path' in locals() and os.path.exists(tmp_path):
                    os.unlink(tmp_path)
            except:
                pass
            raise Exception(f"Error processing Excel file: {str(e)}")

    def _read_sheet_data(self, ws) -> list[list]:
        """
        Read all data from worksheet as a 2D array.
        Handles merged cells and empty cells properly.
        Preserves numeric types (int, float) and converts other types to string.
        
        Args:
            ws: Worksheet object
        """
        # Validate inputs
        if not hasattr(ws, 'max_row'):
            raise TypeError(f"_read_sheet_data expects a worksheet object, got {type(ws)}")
        
        data = []
        max_row = ws.max_row
        max_col = ws.max_column

        # Start from row 1 (1-based)
        for row_idx in range(1, max_row + 1):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = cell.value
                # Preserve numeric types, convert others to string
                if value is not None:
                    if isinstance(value, (int, bool)):
                        # Keep int as int
                        row_data.append(value)
                    elif isinstance(value, float):
                        # Keep float as float
                        row_data.append(value)
                    else:
                        row_data.append(str(value))
                else:
                    row_data.append(None)
            data.append(row_data)

        return data

    def _extract_sheet_styles(self, ws) -> dict[str, dict]:
        """
        Extract styles from Excel sheet.
        Returns a dict mapping cell coordinates (e.g., "A1", "B2") to style objects.
        Only cells with non-default styles are included.
        
        Args:
            ws: Worksheet object
        """
        styles = {}
        max_row = ws.max_row
        max_col = ws.max_column

        # Start from row 1 (1-based)
        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_style = self._extract_cell_style(cell)

                if cell_style:
                    col_letter = self._get_column_letter(col_idx)
                    cell_ref = f"{col_letter}{row_idx}"
                    styles[cell_ref] = cell_style

        return styles

    def _extract_cell_style(self, cell) -> dict | None:
        """Extract style from a single cell."""
        style = {}

        # Font style
        if cell.font:
            font_style = {}
            if cell.font.bold is True:
                font_style["bold"] = True
            if cell.font.italic is True:
                font_style["italic"] = True
            if cell.font.size and cell.font.size != 11:
                font_style["size"] = cell.font.size
            if cell.font.color and cell.font.color.rgb:
                color = cell.font.color.rgb
                if isinstance(color, str) and color not in ('00000000', 'FF000000'):
                    font_style["color"] = color
            if cell.font.underline and cell.font.underline != 'none':
                font_style["underline"] = True
            if cell.font.strike is True:
                font_style["strike"] = True
            if cell.font.name:
                font_style["name"] = cell.font.name

            if font_style:
                style["font"] = font_style

        # Fill (background color)
        if cell.fill:
            fill_style = {}
            if hasattr(cell.fill, 'patternType') and cell.fill.patternType:
                if cell.fill.fgColor and cell.fill.fgColor.rgb:
                    fg_color = cell.fill.fgColor.rgb
                    if isinstance(fg_color, str) and fg_color not in ('00000000', 'FFFFFFFF'):
                        fill_style["fgColor"] = fg_color
                if cell.fill.bgColor and cell.fill.bgColor.rgb:
                    bg_color = cell.fill.bgColor.rgb
                    if isinstance(bg_color, str) and bg_color not in ('00000000', 'FFFFFFFF'):
                        fill_style["bgColor"] = bg_color

            if fill_style:
                style["fill"] = fill_style

        # Alignment
        if cell.alignment:
            align_style = {}
            if cell.alignment.horizontal and cell.alignment.horizontal != 'general':
                align_style["horizontal"] = cell.alignment.horizontal
            if cell.alignment.vertical and cell.alignment.vertical != 'top':
                align_style["vertical"] = cell.alignment.vertical
            if cell.alignment.wrapText:
                align_style["wrapText"] = True

            if align_style:
                style["alignment"] = align_style

        # Border
        if cell.border:
            border_style = {}
            for side in ['top', 'bottom', 'left', 'right']:
                side_attr = getattr(cell.border, side, None)
                if side_attr and side_attr.style:
                    side_dict = {}
                    if side_attr.style:
                        side_dict["style"] = side_attr.style
                    if side_attr.color and side_attr.color.rgb:
                        side_dict["color"] = side_attr.color.rgb
                    if side_dict:
                        border_style[side] = side_dict
            
            if border_style:
                style["border"] = border_style

        # Number format
        if cell.number_format and cell.number_format != 'General':
            style["numberFormat"] = cell.number_format

        # Return style dict only if it has content
        return style if style else None

    def _get_column_letter(self, col_idx: int) -> str:
        """Convert 1-based column index to Excel letter."""
        result = ""
        while col_idx > 0:
            col_idx -= 1
            result = chr(65 + (col_idx % 26)) + result
            col_idx //= 26
        return result

    def _extract_sheet_dimensions(self, ws) -> dict[str, Any]:
        """
        Extract row heights, column widths, and merged cells from Excel sheet.
        
        Args:
            ws: Worksheet object
        """
        dimensions = {}
        
        # Extract row heights
        row_heights = {}
        for row_idx in range(1, ws.max_row + 1):
            if row_idx in ws.row_dimensions:
                height = ws.row_dimensions[row_idx].height
                if height is not None:
                    row_heights[str(row_idx)] = height
        
        if row_heights:
            dimensions["rowHeights"] = row_heights
        
        # Extract column widths
        column_widths = {}
        for col_idx in range(1, ws.max_column + 1):
            col_letter = self._get_column_letter(col_idx)
            if col_letter in ws.column_dimensions:
                width = ws.column_dimensions[col_letter].width
                if width is not None:
                    column_widths[col_letter] = width
        
        if column_widths:
            dimensions["columnWidths"] = column_widths
        
        # Extract merged cells
        if ws.merged_cells:
            merges = []
            for merged_range in ws.merged_cells:
                merges.append(str(merged_range))
            if merges:
                dimensions["merges"] = merges
        
        return dimensions
