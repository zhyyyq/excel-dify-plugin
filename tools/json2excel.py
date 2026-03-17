import json
from collections.abc import Generator
from typing import Any

from dify_plugin import Tool
from dify_plugin.entities.tool import ToolInvokeMessage

import pandas as pd
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.styles.colors import Color


class Json2ExcelTool(Tool):
    """
    Convert JSON to Excel with optional formatting and meta information support.

    Supports:
    - [format] reserved key for row heights and column widths
    - [meta] reserved key for inserting custom content before header
    - [styles] reserved key for cell styling (font, fill, alignment)
    Excel prohibits sheet names containing [ ] characters, so all are safe.
    """

    def _invoke(self, tool_parameters: dict[str, Any]) -> Generator[ToolInvokeMessage]:
        json_str = tool_parameters['json_str']
        
        # Number of rows to skip from the beginning when writing
        skip_rows = int(tool_parameters.get('skip_rows', 0))

        # Parse JSON and extract sheets data, meta, format, and styles configuration
        payload = self._load_json(json_str)
        sheets_data, meta_cfg, format_cfg, styles_cfg = self._extract_sheets_data(payload)
        defaults, sheet_formats, warnings = self._prepare_format_sections(format_cfg, sheets_data.keys())

        # Create Excel file using openpyxl directly
        from openpyxl import Workbook
        wb = Workbook()
        
        # Remove default sheet
        default_sheet = wb.active

        try:
            for sheet_name, data in sheets_data.items():
                # Create new sheet
                ws = wb.create_sheet(title=sheet_name)
                
                # Remove default sheet if this is the first sheet
                if default_sheet:
                    wb.remove(default_sheet)
                    default_sheet = None

                if not isinstance(data, list):
                    raise Exception(f"Value for sheet '{sheet_name}' must be a list.")

                # Skip rows if specified
                if skip_rows > 0 and len(data) > skip_rows:
                    data = data[skip_rows:]

                # Get meta rows for this sheet
                meta_rows = meta_cfg.get(sheet_name, [])
                
                # Write meta rows first (row 1, 2, ...)
                for row_idx, meta_row in enumerate(meta_rows, start=1):
                    if isinstance(meta_row, dict):
                        for col_key, value in meta_row.items():
                            col_letter = self._parse_column_identifier(col_key, f"[meta] row {row_idx}")
                            cell = ws[f"{col_letter}{row_idx}"]
                            cell.value = str(value) if value is not None else ""
                            cell.alignment = Alignment(vertical='center')

                # Write data starting after meta rows
                data_start_row = len(meta_rows) + 1
                
                # Write all rows directly (2D array format)
                for row_offset, row in enumerate(data, start=data_start_row):
                    if isinstance(row, list):
                        for col_idx, value in enumerate(row, start=1):
                            col_letter = get_column_letter(col_idx)
                            cell = ws[f"{col_letter}{row_offset}"]
                            cell.value = str(value) if value is not None else ""
                            cell.alignment = Alignment(vertical='center')

                # Apply formatting if configured
                self._apply_formatting(
                    worksheet=ws,
                    sheet_name=sheet_name,
                    defaults=defaults,
                    sheet_format=sheet_formats.get(sheet_name, {}),
                    data_start_row=data_start_row
                )

                # Apply cell styles if configured
                sheet_styles = styles_cfg.get(sheet_name, {})
                self._apply_cell_styles(ws, sheet_styles, len(meta_rows))

            # Save workbook to buffer
            excel_buffer = BytesIO()
            wb.save(excel_buffer)

        except Exception as e:
            raise Exception(f"Error converting data to Excel: {str(e)}")

        # Create blob message with the Excel bytes
        try:
            excel_buffer.seek(0)
            excel_bytes = excel_buffer.getvalue()
            filename = tool_parameters.get('filename', 'Converted Data')
            filename = f"{filename.replace(' ', '_')}.xlsx"

            # Output warnings if any
            if warnings:
                yield self.create_text_message(f"⚠️ Warning: {warnings}")

            yield self.create_text_message(f"Excel file '{filename}' generated successfully")

            yield self.create_blob_message(
                blob=excel_bytes,
                meta={
                    "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "filename": filename
                }
            )
        except Exception as e:
            raise Exception(f"Error creating Excel file message: {str(e)}")

    def _load_json(self, json_str: str) -> Any:
        """Parse JSON string and return the payload."""
        try:
            return json.loads(json_str)
        except json.JSONDecodeError as exc:
            raise Exception(f"Invalid JSON format: {exc}")

    def _extract_sheets_data(self, payload: Any) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any], dict[str, Any]]:
        """
        Extract sheets data, meta configuration, format configuration, and styles from payload.

        Returns:
            (sheets_data, meta_cfg, format_cfg, styles_cfg) where:
            - sheets_data: dict mapping sheet names to record lists
            - meta_cfg: dict mapping sheet names to meta rows (list of dicts)
            - format_cfg: dict containing format configuration
            - styles_cfg: dict mapping sheet names to cell styles
        """
        # Handle 2D array format: [[header1, header2], [row1], [row2]]
        if isinstance(payload, list) and len(payload) > 0 and isinstance(payload[0], list):
            # Keep 2D array format as-is
            return {"Sheet1": payload}, {}, {}, {}

        if isinstance(payload, list):
            # Single sheet without meta, formatting, or styles (records format)
            return {"Sheet1": payload}, {}, {}, {}

        if isinstance(payload, dict):
            # Extract meta configuration
            if "[meta]" in payload:
                meta_cfg = payload["[meta]"]
                if meta_cfg is None:
                    meta_cfg = {}
                elif isinstance(meta_cfg, dict):
                    # Per-sheet meta: {"Sheet1": [...], "Sheet2": [...]}
                    pass
                elif isinstance(meta_cfg, list):
                    # Global meta applies to first sheet
                    first_sheet = None
                    for k in payload.keys():
                        if k not in ("[meta]", "[format]", "[styles]"):
                            first_sheet = k
                            break
                    if first_sheet:
                        meta_cfg = {first_sheet: meta_cfg}
                    else:
                        meta_cfg = {}
                else:
                    raise Exception("The '[meta]' section must be an array or object when provided.")
                # Validate each sheet's meta is a list
                for sheet_name, meta_rows in meta_cfg.items():
                    if meta_rows is None:
                        meta_cfg[sheet_name] = []
                    elif not isinstance(meta_rows, list):
                        raise Exception(f"The '[meta].{sheet_name}' section must be an array.")
            else:
                meta_cfg = {}

            # Extract format configuration
            if "[format]" in payload:
                format_cfg = payload["[format]"]
                if not isinstance(format_cfg, dict):
                    raise Exception("The '[format]' section must be an object when provided.")
            else:
                format_cfg = {}

            # Extract styles configuration
            if "[styles]" in payload:
                styles_cfg = payload["[styles]"]
                if styles_cfg is None:
                    styles_cfg = {}
                elif isinstance(styles_cfg, dict):
                    # Could be global (cell_ref -> style) or per-sheet
                    # Check if it looks like global styles (keys are cell refs like "A1")
                    # or per-sheet (keys are sheet names)
                    has_cell_refs = any(self._looks_like_cell_ref(k) for k in styles_cfg.keys())
                    if has_cell_refs:
                        # Global styles - apply to first sheet
                        first_sheet = None
                        for k in payload.keys():
                            if k not in ("[meta]", "[format]", "[styles]"):
                                first_sheet = k
                                break
                        if first_sheet:
                            styles_cfg = {first_sheet: styles_cfg}
                        else:
                            styles_cfg = {}
                elif not isinstance(styles_cfg, dict):
                    raise Exception("The '[styles]' section must be an object when provided.")
            else:
                styles_cfg = {}

            # Extract sheets data (all keys except [meta], [format], and [styles])
            sheets = {k: v for k, v in payload.items() if k not in ("[meta]", "[format]", "[styles]")}
            if not sheets:
                raise Exception("At least one sheet must be provided.")

            return sheets, meta_cfg, format_cfg, styles_cfg

        raise Exception("JSON must be an array (single sheet) or object (multiple sheets).")

    def _looks_like_cell_ref(self, key: str) -> bool:
        """Check if a key looks like a cell reference (e.g., 'A1', 'B2')."""
        if not isinstance(key, str) or not key:
            return False
        # Cell refs like A1, B2, AA10, etc.
        if len(key) < 2:
            return False
        # Must start with letters and end with numbers
        letters = ""
        numbers = ""
        for c in key:
            if c.isalpha():
                letters += c
            elif c.isdigit():
                numbers += c
            else:
                return False
        return bool(letters) and bool(numbers)

    def _prepare_format_sections(
        self,
        format_cfg: dict[str, Any],
        sheet_names: set[str]
    ) -> tuple[dict[str, Any], dict[str, Any], str]:
        """
        Validate and prepare format configuration sections.

        Returns:
            (defaults, sheet_formats, warnings) where:
            - defaults: global default formatting
            - sheet_formats: per-sheet formatting overrides
            - warnings: warning message if any (empty string if no warnings)
        """
        if not format_cfg:
            return {}, {}, ""

        # Extract and normalize defaults and sheet_formats
        defaults = format_cfg.get("defaults")
        sheet_formats = format_cfg.get("sheets")

        # Validate types and normalize falsy values to empty dicts
        if "defaults" in format_cfg:
            if defaults is None:
                defaults = {}
            elif not isinstance(defaults, dict):
                raise Exception("The '[format].defaults' section must be an object.")
        else:
            defaults = {}

        if "sheets" in format_cfg:
            if sheet_formats is None:
                sheet_formats = {}
            elif not isinstance(sheet_formats, dict):
                raise Exception("The '[format].sheets' section must be an object.")
        else:
            sheet_formats = {}

        # Check for unknown sheet references (warning mode)
        unknown = set(sheet_formats.keys()) - set(sheet_names)
        warning_msg = ""
        if unknown:
            missing = ", ".join(sorted(unknown))
            warning_msg = f"The '[format].sheets' section references unknown sheets: {missing}. These configurations were ignored."
            # Remove unknown sheets from sheet_formats
            for unknown_sheet in unknown:
                del sheet_formats[unknown_sheet]

        # Validate each sheet format is an object
        for sheet_name, cfg in sheet_formats.items():
            if cfg is None:
                sheet_formats[sheet_name] = {}
            elif not isinstance(cfg, dict):
                raise Exception(f"The '[format].sheets.{sheet_name}' section must be an object.")

        return defaults, sheet_formats, warning_msg

    def _apply_formatting(
        self,
        worksheet,
        sheet_name: str,
        defaults: dict[str, Any],
        sheet_format: dict[str, Any],
        data_start_row: int = 1
    ) -> None:
        """
        Apply formatting to worksheet following priority rules:
        1. defaults.rowHeight/columnWidth (global defaults)
        2. sheets.<name>.rowHeight/columnWidth (sheet defaults)
        3. defaults.rowHeights/columnWidths (global specific rows/columns)
        4. sheets.<name>.rowHeights/columnWidths (sheet specific rows/columns)

        Args:
            worksheet: The openpyxl worksheet object
            sheet_name: Name of the sheet
            defaults: Global default formatting
            sheet_format: Per-sheet formatting overrides
            data_start_row: Row number where data starts (after meta rows, before header)
        """
        max_row = max(worksheet.max_row, 1)
        max_col = max(worksheet.max_column, 1)

        # Apply vertical center alignment to all cells by default
        for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.alignment = Alignment(vertical='center')

        # Apply uniform row heights (priority 1 & 2)
        self._apply_uniform_row_height(
            worksheet, max_row, defaults.get("rowHeight"), "defaults.rowHeight"
        )
        self._apply_uniform_row_height(
            worksheet, max_row, sheet_format.get("rowHeight"), f"sheets.{sheet_name}.rowHeight"
        )

        # Apply uniform column widths (priority 1 & 2)
        self._apply_uniform_column_width(
            worksheet, max_col, defaults.get("columnWidth"), "defaults.columnWidth"
        )
        self._apply_uniform_column_width(
            worksheet, max_col, sheet_format.get("columnWidth"), f"sheets.{sheet_name}.columnWidth"
        )

        # Apply specific row heights (priority 3 & 4)
        self._apply_row_map(
            worksheet, defaults.get("rowHeights"), "defaults.rowHeights"
        )
        self._apply_row_map(
            worksheet, sheet_format.get("rowHeights"), f"sheets.{sheet_name}.rowHeights"
        )

        # Apply specific column widths (priority 3 & 4)
        self._apply_column_map(
            worksheet, defaults.get("columnWidths"), "defaults.columnWidths"
        )
        self._apply_column_map(
            worksheet, sheet_format.get("columnWidths"), f"sheets.{sheet_name}.columnWidths"
        )

    def _apply_meta_content(self, worksheet, meta_rows: list[dict[str, Any]]) -> None:
        """
        Apply meta content (custom rows before header) to worksheet.

        Args:
            worksheet: The openpyxl worksheet object
            meta_rows: List of dicts, each dict maps column letter to value
                       e.g., [{"A": "Title", "C": "Date"}, {"A": "Author"}]
        """
        if not meta_rows:
            return

        for row_idx, meta_row in enumerate(meta_rows, start=1):
            if not isinstance(meta_row, dict):
                raise Exception(f"Each meta row must be an object, but got {type(meta_row).__name__} at row {row_idx}.")

            for col_key, value in meta_row.items():
                column_letter = self._parse_column_identifier(col_key, f"[meta] row {row_idx}")
                cell = worksheet[f"{column_letter}{row_idx}"]
                cell.value = str(value) if value is not None else ""
                cell.alignment = Alignment(vertical='center')

    def _apply_cell_styles(self, worksheet, styles: dict[str, dict], meta_row_count: int = 0) -> None:
        """
        Apply cell styles from [styles] configuration.

        Args:
            worksheet: The openpyxl worksheet object
            styles: Dict mapping cell reference (e.g., "A1", "B2") to style object
            meta_row_count: Number of meta rows inserted before header
        """
        if not styles:
            return

        for cell_ref, style in styles.items():
            if not isinstance(style, dict):
                continue

            try:
                cell = worksheet[cell_ref]
            except (KeyError, ValueError):
                # Invalid cell reference, skip
                continue

            # Apply font styles
            font_cfg = style.get("font")
            if font_cfg and isinstance(font_cfg, dict):
                font_kwargs = {}
                if font_cfg.get("bold"):
                    font_kwargs["bold"] = True
                if font_cfg.get("italic"):
                    font_kwargs["italic"] = True
                if font_cfg.get("size"):
                    font_kwargs["size"] = font_cfg["size"]
                if font_cfg.get("underline"):
                    font_kwargs["underline"] = "single"
                if font_cfg.get("color"):
                    font_kwargs["color"] = font_cfg["color"]

                if font_kwargs:
                    if cell.font:
                        # Merge with existing font
                        existing_font = cell.font
                        cell.font = Font(
                            bold=font_kwargs.get("bold", existing_font.bold),
                            italic=font_kwargs.get("italic", existing_font.italic),
                            size=font_kwargs.get("size", existing_font.size),
                            underline=font_kwargs.get("underline", existing_font.underline),
                            color=font_kwargs.get("color", existing_font.color),
                            name=existing_font.name,
                        )
                    else:
                        cell.font = Font(**font_kwargs)

            # Apply fill styles
            fill_cfg = style.get("fill")
            if fill_cfg and isinstance(fill_cfg, dict):
                fill_kwargs = {"patternType": "solid"}
                if fill_cfg.get("fgColor"):
                    fill_kwargs["fgColor"] = fill_cfg["fgColor"]
                if fill_cfg.get("bgColor"):
                    fill_kwargs["bgColor"] = fill_cfg["bgColor"]

                cell.fill = PatternFill(**fill_kwargs)

            # Apply alignment
            align_cfg = style.get("alignment")
            if align_cfg and isinstance(align_cfg, dict):
                align_kwargs = {}
                if align_cfg.get("horizontal"):
                    align_kwargs["horizontal"] = align_cfg["horizontal"]
                if align_cfg.get("vertical"):
                    align_kwargs["vertical"] = align_cfg["vertical"]

                if align_kwargs:
                    if cell.alignment:
                        # Merge with existing alignment
                        existing = cell.alignment
                        cell.alignment = Alignment(
                            horizontal=align_kwargs.get("horizontal", existing.horizontal),
                            vertical=align_kwargs.get("vertical", existing.vertical),
                        )
                    else:
                        cell.alignment = Alignment(**align_kwargs)

    def _apply_uniform_row_height(self, worksheet, row_count: int, value: Any, label: str) -> None:
        """Apply the same height to all rows."""
        height = self._coerce_positive_number(value, label)
        if height is None:
            return
        for row in range(1, row_count + 1):
            worksheet.row_dimensions[row].height = height

    def _apply_uniform_column_width(self, worksheet, column_count: int, value: Any, label: str) -> None:
        """Apply the same width to all columns."""
        width = self._coerce_positive_number(value, label)
        if width is None:
            return
        for col_idx in range(1, column_count + 1):
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = width

    def _apply_row_map(self, worksheet, mapping: Any, label: str) -> None:
        """Apply specific heights to individual rows."""
        if mapping is None:
            return
        if not isinstance(mapping, dict):
            raise Exception(f"The '{label}' section must be an object.")

        for raw_row, raw_height in mapping.items():
            row_index = self._parse_row_identifier(raw_row, label)
            height = self._coerce_positive_number(raw_height, f"{label}[{raw_row}]")
            worksheet.row_dimensions[row_index].height = height

    def _apply_column_map(self, worksheet, mapping: Any, label: str) -> None:
        """Apply specific widths to individual columns."""
        if mapping is None:
            return
        if not isinstance(mapping, dict):
            raise Exception(f"The '{label}' section must be an object.")

        for raw_col, raw_width in mapping.items():
            column_letter = self._parse_column_identifier(raw_col, label)
            width = self._coerce_positive_number(raw_width, f"{label}[{raw_col}]")
            worksheet.column_dimensions[column_letter].width = width

    def _parse_row_identifier(self, raw_value: Any, label: str) -> int:
        """
        Parse row identifier to 1-based integer.

        Accepts: "1", "2", 1, 2, etc.
        """
        try:
            row = int(str(raw_value))
        except (TypeError, ValueError):
            raise Exception(f"Row identifiers in '{label}' must be positive integers.")

        if row <= 0:
            raise Exception(f"Row identifiers in '{label}' must be 1-based positive integers.")

        return row

    def _parse_column_identifier(self, raw_value: Any, label: str) -> str:
        """
        Parse column identifier to Excel letter format.

        Accepts:
        - Letters: "A", "B", "AA", "AB", etc.
        - 1-based integers: 1, 2, "1", "2", etc.

        Returns: Excel column letter (e.g., "A", "B", "AA")
        """
        if isinstance(raw_value, int):
            index = raw_value
        elif isinstance(raw_value, str):
            token = raw_value.strip()
            if not token:
                raise Exception(f"Column identifiers in '{label}' cannot be empty.")

            if token.isdigit():
                # Numeric string like "1", "2"
                index = int(token)
            elif token.isalpha():
                # Letter string like "A", "B", "AA"
                return token.upper()
            else:
                raise Exception(f"Column identifiers in '{label}' must be letters or integers.")
        else:
            raise Exception(f"Column identifiers in '{label}' must be letters or integers.")

        # Convert 1-based index to Excel letter
        if index <= 0:
            raise Exception(f"Column identifiers in '{label}' must be 1-based positive integers.")

        return get_column_letter(index)

    def _coerce_positive_number(self, value: Any, label: str) -> float | None:
        """
        Coerce value to a positive number.

        Returns None if value is None, otherwise validates and returns float.
        """
        if value is None:
            return None

        try:
            number = float(value)
        except (TypeError, ValueError):
            raise Exception(f"The value for '{label}' must be a positive number.")

        if number <= 0:
            raise Exception(f"The value for '{label}' must be greater than zero.")

        return number
